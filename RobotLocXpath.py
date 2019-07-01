import selenium as se
from selenium import webdriver

import time
import datetime
import json

from openpyxl import Workbook
from openpyxl import load_workbook


# Read excel ID value
def workWithXl(filen):
    wb = load_workbook(filename=filen)
    sheets = wb['Sheet1']
    # print (sheets['A3'].value)
    for row in sheets.rows:
        print(row[0].value)
        row[11].value = 'Done'
    wb.save(filename=filen)


def GetCmdDict(filename):
    cmdStrs = {}

    f = open(filename, "r")
    for row in f:
        # ignore the comment line
        if str(row)[0] == '#':
            # print ('comment {0}'.format(row))
            continue
        elif str(row)[0].lower() == 'r':
            row = row.split("->")
            ky = row[0]
            val = row[1]
            cmdStrs[ky] = val

    f.close()
    return cmdStrs


if __name__ == "__main__":
    tnow = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    print('Start at {0}'.format(tnow))

    # open the URL
    loginCmds = GetCmdDict("Login.prs")
    robotCmds = GetCmdDict("Process.prs")
    brows = webdriver.Chrome()
    # brows.get(loginCmds['URL'])
    # time.sleep(2)

    # Action by the Process.prs
    # Robot Process 1:Login
    brows.find_element_by_xpath(xpath=robotCmds['robot_usrname']).send_keys(loginCmds['username'])
    brows.find_element_by_xpath(xpath=robotCmds['robot_passwd']).send_keys(loginCmds['password'])
    brows.find_element_by_xpath(xpath=robotCmds['robot_login']).click()


    # Read ID Look
    wb = load_workbook(filename='ceb.xlsx')

    # Process 1.1 POS for terminator deletion
    sheetPOS = wb['POS']
    i = 0
    for row in sheetPOS.rows:
        if i == 0:
            i = i + 1
            continue
        elif i > 0:
            # Action 1 - user terminators management
            strID = row[0].value
            # print (strID)
            row[-1].value = 'Del action 1 done'
    # Log operations
    wb.save(filename='ceb.xlsx')

    # Process 1.2 POS for information deletion
    sheetPOS = wb['POS']
    i = 0
    for row in sheetPOS.rows:
        if i == 0:
            i = i + 1
            continue
        elif i > 0:
            # Action 2 - user information management
            strID = row[0].value
            # print (strID)
            row[-2].value = 'Del action 2 done'
    # Log operations
    wb.save(filename='ceb.xlsx')

    # Process 2.1 PAY
    sheetPAY = wb['PAY']
    i = 0
    for row in sheetPAY.rows:
        if i == 0:
            i = i + 1
            continue
        elif i > 0:
            # Action 1 - user terminators management
            strID = row[0].value
            # print(strID)
            row[-1].value = 'Del action 1 done'
    # Log operations
    wb.save(filename='ceb.xlsx')

    # Process 2.2 PAY
    sheetPAY = wb['PAY']
    i = 0
    for row in sheetPAY.rows:
        if i == 0:
            i = i + 1
            continue
        elif i > 0:
            # Action 2 - user information management
            strID = row[0].value
            print(strID)
            row[-2].value = 'Del action 2 done'
    # Log operations
    wb.save(filename='ceb.xlsx')

    wb.close()

    # done
    tnow = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    print('Done {0}'.format(tnow))
